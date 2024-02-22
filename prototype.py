from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl #https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html  documentação openpyxl

from objectsprototype import *

dominio = 'https://bv.fapesp.br/pt/'

def mestrado():

    driver = webdriver.Firefox()
    driver.get(dominio)
    driver.maximize_window()
    time.sleep(2)
    #abre o site da biblioteca virtual

    m = 1
    pagina = PáginaFapesp(driver)

    pagina.fom.click()
    pagina.mest.click()
    pagina.inst.click()
    pagina.a1.click()
    time.sleep(2)
    pagina.a2.click()

    nome_instituto = []
    num_de_bolsas = []
    #listas para colocar informações a planilhar

    pagina.unid1_cebimar.click()
    pagina.refino.click()
    time.sleep(5)
    #abre a primeira unidade
    

    nome_instituto.append(pagina.pega_nome[19:])
    len_anterior = len(pagina.pega_nome[19:])
    time.sleep(1.5)
    if pagina.retangulo.get_attribute('height') == '0':
        text_bolsas = 'Bolsas: 0'
    else:
        pagina.retangulo.click()
        time.sleep(2)
        text_bolsas = pagina.num_bolsas
    num_temp = int(text_bolsas[8:])
    num_de_bolsas.append(num_temp)
    #pega nome e numero de bolsas da primeira unidade


    while m != 58: # itera sobre todas as unidades
        
        m += 1
        time.sleep(3)
        somatorio = num_temp
        
        unid = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,
            '//li[@id = "pivot_expand_UniversidadedeSoPauloUSP"]/ul/li[{}]/input'.format(m))))
        time.sleep(1)
        unid.click()

        time.sleep(1)
        pagina.refino.click()
        time.sleep(1)

        nome_instituto.append(pagina.pega_nome[len_anterior:])
        len_anterior = len(pagina.pega_nome)
        
        time.sleep(1.5)
        if pagina.retangulo.get_attribute('height') == '0':
            text_bolsas = 'Bolsas: 0'
        else:
            pagina.retangulo.click()
            time.sleep(1)
            text_bolsas = pagina.num_bolsas
            time.sleep(1)
        num_temp = int(text_bolsas[8:])

        num_final = num_temp - somatorio    
        num_de_bolsas.append(num_final)
    print(nome_instituto)
    print(num_de_bolsas)


    planilha = openpyxl.Workbook()
    sheet = planilha.active

    sheet['A1'] = 'Unidade'
    sheet['B1'] = 'Bolsas mestrado'

    i = 2
    for elemento in nome_instituto:
        sheet['A{}'.format(i)] = elemento
        i += 1

    j = 2
    for elemento in num_de_bolsas:
        sheet['B{}'.format(j)] = elemento
        j += 1

    planilha.save('Bolsas Fapesp_final.xlsx')

    #escrever na planilha: criar workbook, criar planilha e salvar depois da escrita

#parei na implementação de page objects instantaneamente antes do laço de obtenção da lista do doutorado

def doutorado():

    driver = webdriver.Firefox()
    driver.get(dominio)
    driver.maximize_window()
    time.sleep(2)
    #abre o site da biblioteca virtual

    d = 1
    pagina = PáginaFapesp(driver)
    pagina.fom.click()
    pagina.pag_doc.click()
    pagina.inst.click()
    pagina.a1.click()
    time.sleep(2)
    pagina.a2.click()

    nome_instituto = []
    num_de_bolsas = []
    #listas para colocar informações a planilhar

    pagina.unid1_cebimar.click()
    pagina.refino.click()
    time.sleep(5)
    #abre a primeira unidade

    nome_instituto.append(pagina.pega_nome[19:])
    len_anterior = len(pagina.pega_nome[19:])
    time.sleep(1.5)
    if pagina.retangulo.get_attribute('height') == '0':
        text_bolsas = 'Bolsas: 0'
    else:
        pagina.retangulo.click()
        time.sleep(2)
        text_bolsas = pagina.num_bolsas
    num_temp = int(text_bolsas[8:])
    num_de_bolsas.append(num_temp)
    #pega nome e numero de bolsas da primeira unidade


    while d != 53: # itera sobre todas as unidades

        d += 1
        time.sleep(3)
        somatorio = num_temp
        
        unid = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,
            '//li[@id = "pivot_expand_UniversidadedeSoPauloUSP"]/ul/li[{}]/input'.format(d))))
        time.sleep(1)
        unid.click()

        refino = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,
            '//div[@id="refinamento_instfacet"]/li[2]/input')))
        time.sleep(1)
        refino.click()
        time.sleep(1)

        ni = WebDriverWait(driver,15).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo"]/div[2]/div/div[1]/section/div[1]/span')))
        nome_instituto.append(ni.text[len_anterior + 3:])
        len_anterior = len(ni.text)
        
        retangulo = WebDriverWait(driver,15).until(
            EC.presence_of_element_located((By.XPATH, '//body/div[3]/div/div[6]/section/div/div[1]/*[name()="svg" and @class = "svg-content"]/*[name()="g"]/*[name()="g"][34]/*[local-name()="rect"][2]')))
        time.sleep(1.5)
        verifica_nulo = retangulo.get_attribute('height')
        if verifica_nulo == '0':
            text_bolsas = 'Bolsas: 0'
        else:
            retangulo.click()
            time.sleep(1)
            num_bolsas = WebDriverWait(driver,10).until(
                EC.presence_of_element_located((By.XPATH, '//body/div[@class = "d3-tip n"]/span')))
            text_bolsas = num_bolsas.text
            time.sleep(1)
        num_temp = int(text_bolsas[8:])

        num_final = num_temp - somatorio    
        num_de_bolsas.append(num_final)
        
    print(nome_instituto)
    print(num_de_bolsas)


    nome_instituto = ['Centro de Biologia Marinha CEBIMAR', 'gia Marinha CEBIMAR - Centro de Energia Nuclear na Agricultura CENA', ' - Escola Politécnica EP', ' - Escola Superior de Agricultura Luiz de Queiroz ESALQ', ' - Escola de Artes, Ciências e Humanidades EACH', ' - Escola de Comunicações e Artes ECA', ' - Escola de Educação Física e Esporte EEFE', ' - Escola de Educação Física e Esporte de Ribeirão Preto EEFERP', ' - Escola de Enfermagem EE', ' - Escola de Enfermagem de Ribeirão Preto EERP', ' - Escola de Engenharia de Lorena EEL', ' - Escola de Engenharia de São Carlos EESC', ' - Faculdade de Arquitetura e Urbanismo FAU', ' - Faculdade de Ciências Farmacêuticas FCF', ' - Faculdade de Ciências Farmacêuticas de Ribeirão Preto FCFRP', ' - Faculdade de Direito FD', ' - Faculdade de Economia, Administração e Contabilidade FEA', ' - Faculdade de Economia, Administração e Contabilidade de Ribeirão Preto FEARP', ' - Faculdade de Educação FE', ' - Faculdade de Filosofia, Ciências e Letras de Ribeirão Preto FFCLRP', ' - Faculdade de Filosofia, Letras e Ciências Humanas FFLCH', ' - Faculdade de Medicina FM', ' - Faculdade de Medicina Veterinária e Zootecnia FMVZ', ' - Faculdade de Medicina de Ribeirão Preto FMRP', ' - Faculdade de Odontologia FO', ' - Faculdade de Odontologia de Bauru FOB', ' - Faculdade de Odontologia de Ribeirão Preto FORP', ' - Faculdade de Saúde Pública FSP', ' - Faculdade de Zootecnia e Engenharia de Alimentos FZEA', ' - Hospital de Reabilitação de Anomalias Craniofaciais HRAC', ' - Instituto Oceanográfico IO', ' - Instituto de Arquitetura e Urbanismo de São Carlos IAU', ' - Instituto de Astronomia, Geofísica e Ciências Atmosféricas IAG', ' - Instituto de Biociências IB', ' - Instituto de Ciências Biomédicas ICB', ' - Instituto de Ciências Matemáticas e de Computação ICMC', ' - Instituto de Eletrotécnica e Energia IEE', ' - Instituto de Energia e Ambiente IEE', ' - Instituto de Estudos Brasileiros IEB', ' - Instituto de Física IF', ' - Instituto de Física de São Carlos IFSC', ' - Instituto de Física e Química de São Carlos IFQSC', ' - Instituto de Geociências IGC', ' - Instituto de Matemática e Estatística IME', ' - Instituto de Medicina Tropical de São Paulo IMT', ' - Instituto de Psicologia IP', ' - Instituto de Química IQ', ' - Instituto de Química de São Carlos IQSC', ' - Instituto de Relações Internacionais IRI', ' - Museu de Arqueologia e Etnologia MAE', ' - Museu de Arte Contemporânea MAC', ' - Museu de Zoologia MZ', ' - Pró-Reitoria de Pós-Graduação PRO-PGRAD']
    num_de_bolsas = [0, 8, 10, 27, 2, 5, 1, 2, 0, 1, 1, 7, 8, 6, 8, 1, 0, 0, 1, 10, 18, 5, 4, 16, 3, 5, 3, 9, 9, 0, 1, 5, 2, 7, 21, 8, 0, 1, 0, 8, 1, 0, 4, 3, 0, 0, 5, 3, 2, 2, 1, 1, 0]

    planilha = openpyxl.Workbook()
    sheet = planilha.active

    sheet['A1'] = 'Unidade'
    sheet['B1'] = 'Bolsas doutorado'

    i = 2
    for elemento in nome_instituto:
        sheet['A{}'.format(i)] = elemento
        i += 1

    j = 2
    for elemento in num_de_bolsas:
        sheet['B{}'.format(j)] = elemento
        j += 1

        
    planilha.save('Bolsas Fapesp_doutorado.xlsx')
    #input()

def doutoradodireto():

    dd = 1
    driver = webdriver.Firefox()
    driver.get(dominio)
    driver.maximize_window()
    time.sleep(2)

    pagina = PáginaFapesp(driver)
    pagina.fom.click()
    pagina.pag_doc_dir.click()
    pagina.inst.click()
    pagina.a1.click()
    time.sleep(2)
    pagina.a2.click()
    #abre as instituiçoes de ensino da usp

    nome_instituto = []
    num_de_bolsas = []
    #listas para colocar informações a planilhar

    pagina.unid1_cena.click()
    pagina.refino.click()
    time.sleep(5)
    #abre a primeira unidade  

    nome_instituto.append(pagina.pega_nome[19:])
    len_anterior = len(pagina.pega_nome[19:])
    time.sleep(1.5)
    if pagina.retangulo.get_attribute('height') == '0':
        text_bolsas = 'Bolsas: 0'
    else:
        pagina.retangulo.click()
        time.sleep(2)
        text_bolsas = pagina.num_bolsas
    num_temp = int(text_bolsas[8:])
    num_de_bolsas.append(num_temp)
    #pega nome e numero de bolsas da primeira unidade
    
    while dd != 49: # itera sobre todas as unidades
        dd += 1
        time.sleep(3)
        somatorio = num_temp
        
        unid = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,
            '//li[@id = "pivot_expand_UniversidadedeSoPauloUSP"]/ul/li[{}]/input'.format(dd))))
        time.sleep(1)
        unid.click()
        pagina.refino.click()
        time.sleep(1)

        nome_instituto.append(pagina.pega_nome[len_anterior:])
        len_anterior = len(pagina.pega_nome)
        
        if pagina.retangulo.get_attribute('height') == '0':
            text_bolsas = 'Bolsas: 0'
        else:
            pagina.retangulo.click()
            time.sleep(1)
            text_bolsas = pagina.num_bolsas
            time.sleep(1)
        
        num_temp = int(text_bolsas[8:])
        num_final = num_temp - somatorio    
        num_de_bolsas.append(num_final)
    print(nome_instituto)
    print(num_de_bolsas)
    
    planilha = openpyxl.Workbook()
    sheet = planilha.active

    sheet['A1'] = 'Unidade'
    sheet['B1'] = 'Bolsas doutorado direto'

    i = 2
    for elemento in nome_instituto:
        sheet['A{}'.format(i)] = elemento
        i += 1

    j = 2
    for elemento in num_de_bolsas:
        sheet['B{}'.format(j)] = elemento
        j += 1

    planilha.save('Bolsas Fapesp_doutorado direto.xlsx')

    #['Centro de Energia Nuclear na Agricultura CENA', 'na Agricultura CENA - Centro de Inovação da USP INOVA', ' - Escola Politécnica EP', ' - Escola Superior de Agricultura Luiz de Queiroz ESALQ', ' - Escola de Artes, Ciências e Humanidades EACH', ' - Escola de Comunicações e Artes ECA', ' - Escola de Educação Física e Esporte EEFE', ' - Escola de Educação Física e Esporte de Ribeirão Preto EEFERP', ' - Escola de Enfermagem EE', ' - Escola de Enfermagem de Ribeirão Preto EERP', ' - Escola de Engenharia de Lorena EEL', ' - Escola de Engenharia de São Carlos EESC', ' - Faculdade de Arquitetura e Urbanismo FAU', ' - Faculdade de Ciências Farmacêuticas FCF', ' - Faculdade de Ciências Farmacêuticas de Ribeirão Preto FCFRP', ' - Faculdade de Direito FD', ' - Faculdade de Economia, Administração e Contabilidade FEA', ' - Faculdade de Economia, Administração e Contabilidade de Ribeirão Preto FEARP', ' - Faculdade de Educação FE', ' - Faculdade de Filosofia, Ciências e Letras de Ribeirão Preto FFCLRP', ' - Faculdade de Filosofia, Letras e Ciências Humanas FFLCH', ' - Faculdade de Medicina FM', ' - Faculdade de Medicina Veterinária e Zootecnia FMVZ', ' - Faculdade de Medicina de Ribeirão Preto FMRP', ' - Faculdade de Odontologia FO', ' - Faculdade de Odontologia de Bauru FOB', ' - Faculdade de Odontologia de Ribeirão Preto FORP', ' - Faculdade de Saúde Pública FSP', ' - Faculdade de Zootecnia e Engenharia de Alimentos FZEA', ' - Hospital de Reabilitação de Anomalias Craniofaciais HRAC', ' - Instituto Oceanográfico IO', ' - Instituto de Arquitetura e Urbanismo de São Carlos IAU', ' - Instituto de Astronomia, Geofísica e Ciências Atmosféricas IAG', ' - Instituto de Biociências IB', ' - Instituto de Ciências Biomédicas ICB', ' - Instituto de Ciências Matemáticas e de Computação ICMC', ' - Instituto de Energia e Ambiente IEE', ' - Instituto de Física IF', ' - Instituto de Física de São Carlos IFSC', '', ' - Instituto de Geociências IGC - Instituto de Matemática e Estatística IME', ' - Instituto de Medicina Tropical de São Paulo IMT', ' - Instituto de Psicologia IP', ' - Instituto de Química IQ', ' - Instituto de Química de São Carlos IQSC', ' - Instituto de Relações Internacionais IRI', ' - Museu de Arqueologia e Etnologia MAE', ' - Museu de Arte Contemporânea MAC', ' - Museu de Zoologia MZ']

def pos_doutorado():   

    pd = 1

    driver = webdriver.Firefox()
    driver.get(dominio)
    driver.maximize_window()
    time.sleep(2)

    pagina = PáginaFapesp(driver)
    pagina.fom.click()
    pagina.pag_pos_doc.click()
    pagina.inst.click()
    pagina.a1.click()
    time.sleep(2)
    pagina.a2.click()
    #abre as instituiçoes de ensino da usp

    nome_instituto = []
    num_de_bolsas = []
    #listas para colocar informações a planilhar

    pagina.unid1_cebimar.click()
    pagina.refino.click()
    time.sleep(5)
    #abre a primeira unidade

    nome_instituto.append(pagina.pega_nome[19:])
    len_anterior = len(pagina.pega_nome[19:])
    time.sleep(1.5)
    if pagina.retangulo.get_attribute('height') == '0':
        text_bolsas = 'Bolsas: 0'
    else:
        pagina.retangulo.click()
        time.sleep(2)
        text_bolsas = pagina.num_bolsas
    num_temp = int(text_bolsas[8:])
    num_de_bolsas.append(num_temp)
    #pega nome e numero de bolsas da primeira unidade

    while pd != 62: # itera sobre todas as unidades
    
        pd += 1
        time.sleep(3)
        somatorio = num_temp
        
        unid = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.XPATH,
            '//li[@id = "pivot_expand_UniversidadedeSoPauloUSP"]/ul/li[{}]/input'.format(pd))))
        time.sleep(1)
        unid.click()
        pagina.refino.click()
        time.sleep(1)

        nome_instituto.append(pagina.pega_nome[len_anterior:])
        len_anterior = len(pagina.pega_nome)
        
        if pagina.retangulo.get_attribute('height') == '0':
            text_bolsas = 'Bolsas: 0'
        else:
            pagina.retangulo.click()
            time.sleep(1)
            text_bolsas = pagina.num_bolsas
            time.sleep(1)
        
        num_temp = int(text_bolsas[8:])
        num_final = num_temp - somatorio    
        num_de_bolsas.append(num_final)
        #['Centro de Biologia Marinha CEBIMAR', 'gia Marinha CEBIMAR - Centro de Energia Nuclear na Agricultura CENA', ' - Centro de Inovação da USP INOVA', ' - Escola Politécnica EP', ' - Escola Superior de Agricultura Luiz de Queiroz ESALQ', ' - Escola de Artes, Ciências e Humanidades EACH', ' - Escola de Comunicações e Artes ECA', ' - Escola de Educação Física e Esporte EEFE', ' - Escola de Educação Física e Esporte de Ribeirão Preto EEFERP', ' - Escola de Enfermagem EE', ' - Escola de Enfermagem de Ribeirão Preto EERP', ' - Escola de Engenharia de Lorena EEL', ' - Escola de Engenharia de São Carlos EESC', ' - Faculdade de Arquitetura e Urbanismo FAU', ' - Faculdade de Ciências Farmacêuticas FCF', '', ' - Faculdade de Ciências Farmacêuticas de Ribeirão Preto FCFRP - Faculdade de Direito FD', ' - Faculdade de Direito de Ribeirão Preto FDRP', ' - Faculdade de Economia, Administração e Contabilidade FEA', ' - Faculdade de Economia, Administração e Contabilidade de Ribeirão Preto FEARP', ' - Faculdade de Educação FE', ' - Faculdade de Filosofia, Ciências e Letras de Ribeirão Preto FFCLRP', ' - Faculdade de Filosofia, Letras e Ciências Humanas FFLCH', ' - Faculdade de Medicina FM', ' - Faculdade de Medicina Veterinária e Zootecnia FMVZ', ' - Faculdade de Medicina de Ribeirão Preto FMRP', '', ' - Faculdade de Odontologia FO - Faculdade de Odontologia de Bauru FOB', ' - Faculdade de Odontologia de Ribeirão Preto FORP', ' - Faculdade de Saúde Pública FSP', ' - Faculdade de Zootecnia e Engenharia de Alimentos FZEA', ' - Hospital Universitário HU', ' - Hospital de Reabilitação de Anomalias Craniofaciais HRAC', ' - Instituto Oceanográfico IO', ' - Instituto de Arquitetura e Urbanismo de São Carlos IAU', ' - Instituto de Astronomia, Geofísica e Ciências Atmosféricas IAG', ' - Instituto de Biociências IB', ' - Instituto de Ciências Biomédicas ICB', ' - Instituto de Ciências Matemáticas e de Computação ICMC', ' - Instituto de Eletrotécnica e Energia IEE', ' - Instituto de Energia e Ambiente IEE', ' - Instituto de Estudos Avançados IEA', ' - Instituto de Estudos Brasileiros IEB', ' - Instituto de Física IF', ' - Instituto de Física de São Carlos IFSC', ' - Instituto de Física e Química de São Carlos IFQSC', ' - Instituto de Geociências IGC', ' - Instituto de Matemática e Estatística IME', ' - Instituto de Medicina Tropical de São Paulo IMT', ' - Instituto de Psicologia IP', ' - Instituto de Química IQ', ' - Instituto de Química de São Carlos IQSC', ' - Instituto de Relações Internacionais IRI', ' - Museu Paulista MP', ' - Museu de Arqueologia e Etnologia MAE', ' - Museu de Arte Contemporânea MAC', ' - Museu de Zoologia MZ', ' - Plataforma Científica Pasteur', ' - Pró-Reitoria de Pesquisa PRO-PESQ', ' - Pró-Reitoria de Pós-Graduação PRO-PGRAD', ' - Reitoria RUSP', ' - Sistema Integrado de Bibliotecas SIBi']
    print(nome_instituto)
    print(num_de_bolsas)

    planilha = openpyxl.Workbook()
    sheet = planilha.active

    sheet['A1'] = 'Unidade'
    sheet['B1'] = 'Bolsas pós-doutorado'

    i = 2
    for elemento in nome_instituto:
        sheet['A{}'.format(i)] = elemento
        i += 1

    j = 2
    for elemento in num_de_bolsas:
        sheet['B{}'.format(j)] = elemento
        j += 1

    planilha.save('Bolsas Fapesp_pósdoutorado.xlsx')


if __name__ == '__main__':
    #mestrado()
    #doutorado()
    #doutoradodireto()
    pos_doutorado()
   

